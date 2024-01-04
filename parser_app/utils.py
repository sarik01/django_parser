import re
from itertools import zip_longest
from pprint import pprint

import xlrd
from django.db import connection
from django.db.models import Sum
from openpyxl import load_workbook
from .models import *


#
# def process_excel_and_create_insert(file_path, table_name):
#     df = pd.read_excel(file_path)
#
#     create_table_query = f"CREATE TABLE IF NOT EXISTS {table_name} ("
#     create_table_query += "\n    id SERIAL PRIMARY KEY,"  # Добавление первичного ключа
#
#     for column_name in df.columns:
#         x = re.findall(r"foreignkey|varchar|float|text|boolean", column_name)
#         print(x)
#         if column_name.lower() == 'id':
#             continue  # Пропустить, если столбец с именем 'id' уже существует
#
#         if ('foreignkey' in x):
#             column_name = column_name.replace(' (foreignkey)', '')
#             sql_data_type = 'INTEGER'
#
#         elif ('varchar' in x):
#             column_name = column_name.replace(' (varchar)', '')
#             sql_data_type = 'VARCHAR(255)'
#
#         elif ('float' in x):
#             column_name = column_name.replace(' (float)', '')
#             sql_data_type = 'TEXT'
#
#         elif ('text' in x):
#             column_name = column_name.replace(' (text)', '')
#             sql_data_type = 'TEXT'
#
#         elif ('boolean' in x):
#             column_name = column_name.replace(' (boolean)', '')
#             sql_data_type = 'BOOLEAN'
#
#         create_table_query += f"\n    {column_name} {sql_data_type},"
#
#     for column in df.columns.values:
#         try:
#             column_name, option = column.split(' ')
#             option: str = option.replace('(', '').replace(')', '')
#             if 'foreignkey' == option:
#                 referenced_table = column_name.replace('_id', 's')
#                 create_table_query += f"\n    FOREIGN KEY ({column_name}) REFERENCES {referenced_table}(id),"
#         except:
#             continue
#     create_table_query = create_table_query.rstrip(',') + "\n);"
#     print(create_table_query)
#
#     with connection.cursor() as cursor:
#         cursor.execute(create_table_query)

# # Формирование SQL-запроса для вставки данных
# insert_query = f"INSERT INTO {table_name} ({', '.join(df.columns)}) VALUES "
#
# # Создание строки с заполнителями для значений
# value_placeholders = ', '.join(['%s' for _ in df.columns])
#
# # Создание строки с множеством значений для вставки
# values_str = ', '.join([f"({', '.join(['%s' for _ in df.columns])})" for _ in range(len(df))])
#
# # Сбор итогового SQL-запроса
# insert_query += f"{values_str};"
#
# # Выполнение SQL-запроса для вставки данных
# with connection.cursor() as cursor:
#     cursor.execute(insert_query, df.values.flatten())


def get_comp_ings():
    compound_ingredients = CompoundIngredients.objects.all().select_related('measurement_unit',
                                                                            'group').prefetch_related('ingredients')

    all_data = []

    for compound_ingredient in compound_ingredients:
        ingredients = [x.format() for x in
                       compound_ingredient.ingredients.all().select_related('measurement_unit').order_by('id')]
        add_tables = compound_ingredient.additionaltableci_set.all().order_by('ingredients__id')
        response_data = {
            'title': compound_ingredient.title,
            'measurement_unit': compound_ingredient.measurement_unit.name,
            'group': compound_ingredient.group.name,
            'ingredients': ingredients,
            'total': add_tables.aggregate(total_sales=Sum('sum'))['total_sales']
        }

        all_data.append(response_data)

        # pprint(all_data)
    return all_data

def comp_ing_import(file):
    # paring from excel
    wb = load_workbook(file, read_only=True)
    sheet = wb.active
    lst = []

    for row in sheet.iter_rows(values_only=True):
        filtered_list = list(item for item in row if item is not None)
        # print(filtered_list)
        if filtered_list[3:]:
            x = re.findall(r"Costs shown as for", str(filtered_list[0]))
            if "Costs shown as for" in x:
                continue
        else:
            continue
        lst.append(filtered_list)
    headers = []
    result_list = []
    for sublist in lst:
        row_dict = {}
        if '#' in sublist:
            sublist = [i.lower().replace(' ', '_') for i in sublist]
            headers.append(sublist)
            continue
        for i, value in enumerate(sublist):
            if len(sublist) == len(headers[0]):
                row_dict[headers[0][i]] = value

            else:
                row_dict[headers[1][i]] = value
        del row_dict['#']
        if len(row_dict) + 1 == len(headers[0]):
            row_dict['ingredients'] = []
            result_list.append(row_dict)
        else:
            result_list[-1]['ingredients'].append(row_dict)
    pprint(result_list)
    # save data to db
    comp_ing = {}

    for obj in result_list:
        for k, v in obj.items():
            if k == 'group':
                group, created = Groups.objects.get_or_create(name=v)
                comp_ing['group'] = group
            elif k == 'measurement_unit':
                unit, created = MeasureUnits.objects.get_or_create(name=v)
                comp_ing['measurement_unit'] = unit
            elif k == 'title':
                comp_ing['title'] = v
            elif k == 'ingredients':
                comp_ing['ingredients'] = []
                comp_ing['add_table'] = []
                for ingredient in v:
                    try:
                        ingredient['the_price_of_the_ingredient_by_the_invoice'] = float(
                            ingredient['the_price_of_the_ingredient_by_the_invoice'].replace(' ', '').replace(',', '.'))
                        ingredient['sum'] = float(ingredient['sum'].replace(' ', '').replace(',', '.'))
                    except AttributeError:
                        ingredient['the_price_of_the_ingredient_by_the_invoice'] = float(
                            ingredient['the_price_of_the_ingredient_by_the_invoice'])
                        ingredient['sum'] = float(ingredient['sum'])

                    unit, created = MeasureUnits.objects.get_or_create(name=ingredient['measurement_unit'])
                    ingredient_table, created = Ingredients.objects.get_or_create(name=ingredient['ingredient'],
                                                                                  measurement_unit=unit,
                                                                                  price=ingredient[
                                                                                      'the_price_of_the_ingredient_by_the_invoice'])
                    add_table, created = AdditionalTableCI.objects.get_or_create(quantity=ingredient['quantity'],
                                                                                 calculation_qty=ingredient[
                                                                                     'calculation_qty'],
                                                                                 sum=ingredient['sum'])

                    add_table.ingredients.add(ingredient_table)

                    comp_ing['ingredients'].append(ingredient_table)
                    # print(ingredient['sum'])

                    comp_ing['add_table'].append(add_table)
        # print(comp_ing)

        ci, created = CompoundIngredients.objects.get_or_create(title=comp_ing['title'],
                                                                measurement_unit=comp_ing['measurement_unit'],
                                                                group=comp_ing['group'])
        ci.ingredients.add(*comp_ing['ingredients'])
        ci.save()

        for i in comp_ing['add_table']:
            i.compound_ingredients.add(ci)
            i.save()


    all_data = get_comp_ings()


    print(len(connection.queries))
    return all_data


def upload_file3(file):
    wb = load_workbook(file, read_only=True)
    sheet = wb.active
    lst = []

    for row in sheet.iter_rows(values_only=True):
        filtered_list = list(item for item in row[1:])
        # print(filtered_list)
        if not all(value is None for value in filtered_list):
            if filtered_list:
                x = re.findall(r"Costs shown as for|INVOICE|Details|Total:",
                               str(filtered_list[0]) + str(filtered_list[-4]))
                # print(x)
                if "Costs shown as for" in x or 'INVOICE' in x or 'Details' in x or 'Total:' in x:
                    continue
            else:
                continue
        else:
            continue

        lst.append(filtered_list)

    # Transform data into the desired format
    result = []

    invoice = {}
    ingredient_section = False
    ingredient_header = None  # Initialize ingredient_header

    for row in lst:
        if row[0] == 'Date':
            continue  # Skip header rows

        if row[1] == 'Ingredient':
            ingredient_section = True
            ingredient_header = row
            continue  # Skip ingredient section header row

        if not ingredient_section:
            invoice = {
                'amount': row[-4] if row[-4] != 'none' else None,
                'date': row[1],
                'description': row[-1] if row[-1] != 'none' else None,
                'for_payment': row[-5] if row[-5] != 'none' else None,
                'ingredients': [],  # Initialize an empty list for ingredients
                'paid': row[4] if row[4] != 'none' else None,
                'responsible': row[-3] if row[-3] != 'none' else None,
                'rough_copy': row[-2] if row[-2] != 'none' else None,
                'storage': row[2],
                'supplier': row[3]
            }
            ingredient_section = False  # Reset the flag after processing main invoice row
        else:
            ingredient = {ingredient_header[i]: row[i] for i in range(len(ingredient_header))}
            invoice['ingredients'].append(ingredient)

    result.append(invoice)

    pprint(result)

def without_relations(file):

    # paring from excel
    wb = load_workbook(file, read_only=True)
    sheet = wb.active
    lst = []
    headers = []

    for row in sheet.iter_rows(values_only=True):
        filtered_list = list(item for item in row[1:])
        # print(filtered_list)
        if not all(value is None for value in filtered_list):
            if filtered_list:
                x = re.findall(r"Costs shown as for|INVOICE|Details|Total:|CANCELLATION",
                               str(filtered_list[0]) + str(filtered_list[-4]))
                # print(x)
                if "Costs shown as for" in x or 'INVOICE' in x or 'Details' in x or 'Total:' in x \
                        or 'CANCELLATION' in x:
                    continue
            else:
                continue
        else:
            continue
        lst.append(filtered_list)
    # print(lst)
    result_list = []
    for sublist in lst:
        row_dict = {}
        if '#' in sublist:
            headers = sublist
            continue
        # print(sublist)
        headers = [i for i in headers if i is not None]
        sublist = [i for i in sublist if i is not None]
        # print(headers)
        row_dict = dict(zip_longest(headers, sublist, fillvalue=None))
        # print(row_dict)
        del row_dict['#']

        if len(row_dict) + 1 == len(headers):
            row_dict['relations'] = []
            result_list.append(row_dict)
        else:
            result_list[-1]['relations'].append(row_dict)
    pprint(result_list)


def invoice_import(file):
    # paring from excel
    wb = load_workbook(file, read_only=True)
    sheet = wb.active
    lst = []
    headers = []

    for row in sheet.iter_rows(values_only=True):
        filtered_list = list(item for item in row[1:])
        # print(filtered_list)
        if not all(value is None for value in filtered_list):
            if filtered_list:
                x = re.findall(r"Costs shown as for|INVOICE|Details|Total:|CANCELLATION",
                               str(filtered_list[0]) + str(filtered_list[-4]))
                # print(x)
                if "Costs shown as for" in x or 'INVOICE' in x or 'Details' in x or 'Total:' in x\
                        or 'CANCELLATION' in x:
                    continue
            else:
                continue
        else:
            continue
        lst.append(filtered_list)
    # print(lst)
    result_list = []
    for sublist in lst:
        row_dict = {}
        if '#' in sublist:
            headers = sublist
            continue
        # print(sublist)
        headers = [i for i in headers if i is not None]
        sublist = [i for i in sublist if i is not None]
        # print(headers)
        row_dict = dict(zip_longest(headers, sublist, fillvalue=None))
        # print(row_dict)
        del row_dict['#']

        # if len(row_dict) + 1 == len(headers):
        #     row_dict['relations'] = []
        #     result_list.append(row_dict)
        # else:
        #     result_list[-1]['relations'].append(row_dict)

        if 'Date' in headers:
            row_dict['relations'] = []
            result_list.append(row_dict)
        else:
            result_list[-1]['relations'].append(row_dict)

        try:
            del row_dict[None]
        except KeyError:
            pass
    pprint(result_list)


def write_offs(file):
    # paring from excel
    wb = load_workbook(file, read_only=True)
    sheet = wb.active
    lst = []
    headers = []

    for row in sheet.iter_rows(values_only=True):
        filtered_list = list(item for item in row[1:])
        # print(filtered_list)
        if not all(value is None for value in filtered_list):
            if filtered_list:
                x = re.findall(r"Costs shown as for|INVOICE|Details|Total:|CANCELLATION",
                               str(filtered_list[0]) + str(filtered_list[-4]))
                # print(x)
                if "Costs shown as for" in x or 'INVOICE' in x or 'Details' in x or 'Total:' in x\
                        or 'CANCELLATION' in x:
                    continue
            else:
                continue
        else:
            continue
        lst.append(filtered_list)
    # print(lst)
    result_list = []
    for sublist in lst:
        if '#' in sublist:
            headers = sublist
            continue
        print(sublist)
        headers = [i for i in headers if i is not None]
        sublist = [i for i in sublist if i is not None]
        print(headers)
        row_dict = dict(zip_longest(headers, sublist, fillvalue=None))
        print(row_dict)
        del row_dict['#']

        if 'Date' in headers:
            row_dict['ingredients'] = []
            result_list.append(row_dict)
        else:
            result_list[-1]['ingredients'].append(row_dict)

        try:
            del row_dict[None]
        except KeyError:
            pass

    pprint(result_list)